# main.py - API de Diagnóstico (V6.7) - Corrección Final de Serialización de Fecha

import io
import os
import uuid
import numpy as np
import tensorflow as tf
import firebase_admin
import joblib
import requests
import json
from firebase_admin import credentials, firestore, storage
from datetime import datetime, timezone
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from PIL import Image
from sklearn.preprocessing import LabelEncoder
from typing import List
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# --- INICIALIZACIÓN DE LA APLICACIÓN ---
print("--- Iniciando aplicación FastAPI (V6.7) ---")
app = FastAPI(
    title="API de Diagnóstico de Piezas Hidráulicas",
    description="API final con diagnóstico, persistencia y gestión de reportes.",
    version="FINAL"
)

# --- INICIALIZACIÓN DE FIREBASE ---
try:
    credentials_json_str = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS_JSON')
    if credentials_json_str:
        print("-> Inicializando Firebase desde variable de entorno (modo nube)...")
        credentials_dict = json.loads(credentials_json_str)
        cred = credentials.Certificate(credentials_dict)
    else:
        print("-> Inicializando Firebase desde archivo local (modo desarrollo)...")
        KEY_FILENAME = "serviceAccountKey.json"
        if not os.path.exists(KEY_FILENAME):
            raise FileNotFoundError(f"El archivo de credenciales '{KEY_FILENAME}' no se encontró.")
        cred = credentials.Certificate(KEY_FILENAME)

    firebase_admin.initialize_app(cred, {
        'storageBucket': 'diagnostico-piezas-myd.firebasestorage.app'
    })
    db = firestore.client()
    bucket = storage.bucket()
    print("-> Conexión con Firestore y Storage establecida exitosamente.")
except Exception as e:
    print(f"¡Error Crítico al conectar con Firebase! Error: {e}")
    db = None
    bucket = None

# --- CARGA DE ARTEFACTOS ---
MODEL_FILENAME = "modelo_diagnostico_v1.keras"
PIEZA_ENCODER_FILENAME = "pieza_encoder.joblib"
ESTADO_ENCODER_FILENAME = "estado_encoder.joblib"

required_files = [MODEL_FILENAME, PIEZA_ENCODER_FILENAME, ESTADO_ENCODER_FILENAME]
for filename in required_files:
    if not os.path.exists(filename):
        raise RuntimeError(f"¡Error Crítico! No se encontró el archivo de artefacto '{filename}'.")

model = tf.keras.models.load_model(MODEL_FILENAME)
pieza_encoder = joblib.load(PIEZA_ENCODER_FILENAME)
estado_encoder = joblib.load(ESTADO_ENCODER_FILENAME)
print("-> Modelo y traductores cargados exitosamente.")

# --- CONFIGURACIÓN GLOBAL Y FUNCIONES AUXILIARES ---
IMG_SIZE = (160, 160)
ESTADO_PRIORIDAD = {estado: i for i, estado in enumerate(['optimo', 'desgaste', 'corrosion', 'ruptura'])}

def preprocesar_imagen(image_bytes: bytes):
    img = Image.open(io.BytesIO(image_bytes)).convert('RGB')
    img_resized = img.resize(IMG_SIZE)
    img_array = tf.keras.preprocessing.image.img_to_array(img_resized)
    img_expanded = np.expand_dims(img_array, axis=0)
    img_preprocessed = tf.keras.applications.mobilenet_v2.preprocess_input(img_expanded)
    return img_preprocessed

def generar_sugerencia(estado: str) -> str:
    sugerencias = {
        "ruptura": "¡ALERTA MÁXIMA! Pieza presenta rupturas , pudo haber sido causado por impactos de sólidos o sobrepresión; reemplace el impulsor e investigue la causa raíz en el sistema",
        "corrosion": "Atención: La pieza presenta corrosion , pudo haber sido causado por ataque químico del fluido; reemplace por un material compatible y analice el fluido. La corrosión puede comprometer la integridad estructural...",
        "desgaste": "La pieza muestra desgaste , pudo haber sido por abrasión de partículas o por cavitación; reemplace y corrija el sistema (filtrado o condiciones de succión).",
        "optimo": "Condición ideal. La pieza está en perfecto estado..."
    }
    return sugerencias.get(estado, "No se ha podido determinar una sugerencia clara.")

def guardar_reporte(reporte_data: dict, reporte_id: str):
    if db is None: return
    try:
        reportes_ref = db.collection('reportes')
        reportes_ref.document(reporte_id).set(reporte_data)
        print(f"-> Reporte {reporte_id} guardado en Firestore.")
    except Exception as e:
        print(f"Error al guardar el reporte en Firestore: {e}")

# --- ENDPOINTS DE LA API ---

@app.get("/", summary="Endpoint raíz de bienvenida")
def read_root():
    return {"mensaje": "Bienvenido a la API de Diagnóstico. Vaya a /docs."}

@app.post("/diagnosticar/", summary="Diagnosticar pieza desde 5 vistas y guardar imágenes")
async def diagnosticar_pieza_multivista(files: List[UploadFile] = File(..., description="Cargue exactamente 5 imágenes de la pieza.")):
    if bucket is None: raise HTTPException(status_code=503, detail="Servicio de Storage no disponible.")
    if len(files) != 5: raise HTTPException(status_code=400, detail="Se requieren exactamente 5 imágenes.")
    
    reporte_id = str(uuid.uuid4())
    vectores_prob_pieza, vectores_prob_estado, urls_imagenes = [], [], []

    for i, file in enumerate(files):
        image_bytes = await file.read()
        try:
            filename = f"reportes/{reporte_id}/imagen_{i+1}{os.path.splitext(file.filename)[1]}"
            blob = bucket.blob(filename)
            blob.upload_from_string(image_bytes, content_type=file.content_type)
            blob.make_public()
            urls_imagenes.append(blob.public_url)
            print(f"-> Imagen {i+1} subida a Storage.")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error al subir la imagen {i+1}: {e}")
        try:
            processed_image = preprocesar_imagen(image_bytes)
            predictions = model.predict(processed_image)
            vectores_prob_pieza.append(predictions[0][0])
            vectores_prob_estado.append(predictions[1][0])
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error al procesar '{file.filename}': {str(e)}")
    
    id_pieza_final = np.bincount([np.argmax(p) for p in vectores_prob_pieza]).argmax()
    label_pieza_final = pieza_encoder.inverse_transform([id_pieza_final])[0]
    estados_predichos = [estado_encoder.inverse_transform([np.argmax(p)])[0] for p in vectores_prob_estado]
    estado_final = max(estados_predichos, key=lambda e: ESTADO_PRIORIDAD.get(e, -1))
    prob_estado_promedio = np.mean(vectores_prob_estado, axis=0)
    panel_confianza = {estado_encoder.classes_[i]: f"{p*100:.2f}%" for i, p in enumerate(prob_estado_promedio)}
    
    reporte = {
        "reporte_id": reporte_id,
        "timestamp": datetime.now(timezone.utc),
        "pieza_identificada": label_pieza_final,
        "diagnostico_agregado": estado_final,
        "panel_confianza_estado": panel_confianza,
        "sugerencia": generar_sugerencia(estado_final),
        "diagnosticos_individuales": estados_predichos,
        "urls_imagenes": urls_imagenes
    }
    
    guardar_reporte(reporte, reporte_id)
    return reporte

@app.get("/reportes/", summary="Listar todos los reportes guardados")
async def listar_reportes():
    if db is None: raise HTTPException(status_code=503, detail="Servicio de base de datos no disponible.")
    try:
        reportes_ref = db.collection('reportes').stream()
        reportes = []
        for doc in reportes_ref:
            reporte = doc.to_dict()
            reporte['reporte_id'] = doc.id
            
            # *** CORRECCIÓN CRUCIAL AQUÍ ***
            # Si el campo 'timestamp' existe, lo convertimos a un formato de texto estándar (ISO 8601)
            # antes de enviarlo. Esto evita el error de serialización.
            if reporte.get('timestamp'):
                reporte['timestamp'] = reporte['timestamp'].isoformat()
            
            reportes.append(reporte)
        return reportes
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener reportes: {e}")

@app.delete("/reportes/{reporte_id}", summary="Eliminar un reporte específico")
async def eliminar_reporte(reporte_id: str):
    if db is None: raise HTTPException(status_code=503, detail="Servicio de base de datos no disponible.")
    try:
        blobs = bucket.list_blobs(prefix=f"reportes/{reporte_id}/")
        for blob in blobs:
            blob.delete()
            print(f"-> Imagen {blob.name} eliminada de Storage.")
        
        reporte_ref = db.collection('reportes').document(reporte_id)
        if not reporte_ref.get().exists:
            raise HTTPException(status_code=404, detail="Reporte no encontrado.")
        reporte_ref.delete()
        return {"mensaje": f"Reporte {reporte_id} y sus imágenes han sido eliminados."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar el reporte: {e}")

@app.get("/reportes/{reporte_id}/excel", summary="Descargar reporte profesional en Excel")
async def descargar_excel(reporte_id: str):
    if db is None: raise HTTPException(status_code=503, detail="Servicio de base de datos no disponible.")
    
    doc_ref = db.collection('reportes').document(reporte_id)
    reporte = doc_ref.get().to_dict()
    if not reporte: raise HTTPException(status_code=404, detail="Reporte no encontrado.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Diagnóstico"
    
    font_titulo = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    font_header = Font(name='Calibri', size=12, bold=True)
    alineacion_centro = Alignment(horizontal='center', vertical='center')

    ws['A1'] = "Reporte de Diagnóstico de Pieza Hidráulica"
    ws.merge_cells('A1:E1')
    ws['A1'].font = font_titulo
    ws['A1'].alignment = alineacion_centro
    ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    ws.row_dimensions[1].height = 30

    # *** CORRECCIÓN CRUCIAL AQUÍ TAMBIÉN ***
    # Nos aseguramos de manejar la fecha correctamente, ya sea que venga como objeto o como texto.
    timestamp_obj = reporte.get('timestamp')
    if isinstance(timestamp_obj, datetime):
        fecha_str = timestamp_obj.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(timestamp_obj, str):
        fecha_str = timestamp_obj
    else:
        fecha_str = "N/A"
        
    ws['A3'] = "ID del Reporte:"; ws['B3'] = reporte_id
    ws['A4'] = "Fecha (UTC):"; ws['B4'] = fecha_str
    
    ws['A6'] = "RESULTADOS DEL DIAGNÓSTICO"; ws['A6'].font = font_header
    ws['A7'] = "Pieza Identificada:"; ws['B7'] = reporte.get('pieza_identificada')
    ws['A8'] = "Diagnóstico Agregado Final:"; ws['B8'] = reporte.get('diagnostico_agregado').upper()
    ws['A8'].font = font_header; ws['B8'].font = font_header
    ws['A9'] = "Sugerencia de Acción:"; ws['B9'] = reporte.get('sugerencia')
    
    ws['D6'] = "PANEL DE CONFIANZA"; ws['D6'].font = font_header
    row = 7
    for estado, conf in reporte.get('panel_confianza_estado', {}).items():
        ws[f'D{row}'] = estado.capitalize() + ":"; ws[f'E{row}'] = conf
        row += 1
        
    ws['A13'] = "DIAGNÓSTICO POR VISTA"; ws['A13'].font = font_header
    diagnosticos_ind = reporte.get('diagnosticos_individuales', [])
    for i, diag in enumerate(diagnosticos_ind):
        ws[f'A{14+i}'] = f"Vista {i+1}:"; ws[f'B{14+i}'] = diag

    ws['A20'] = "EVIDENCIA FOTOGRÁFICA"; ws['A20'].font = font_header
    urls = reporte.get("urls_imagenes", [])
    if urls:
        for i, url in enumerate(urls):
            ws[f"{get_column_letter(i + 1)}21"] = f"Vista {i+1}"
            ws[f"{get_column_letter(i + 1)}21"].font = font_header
            ws[f"{get_column_letter(i + 1)}21"].alignment = alineacion_centro
            try:
                response = requests.get(url, stream=True)
                response.raise_for_status()
                image_stream = io.BytesIO(response.content)
                img = OpenpyxlImage(image_stream)
                img.width, img.height = 120, 120
                ws.add_image(img, f"{get_column_letter(i + 1)}22")
                ws.row_dimensions[22].height = 95
            except Exception as e:
                print(f"No se pudo descargar o insertar la imagen {url}. Error: {e}")
                ws[f"{get_column_letter(i + 1)}22"] = "Imagen no disponible"

    ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 30
    ws.column_dimensions['D'].width = 15; ws.column_dimensions['E'].width = 15

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    return StreamingResponse(
        virtual_workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_{reporte_id}.xlsx"}
    )